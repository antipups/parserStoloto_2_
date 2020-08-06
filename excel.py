import zipfile
from os import path, remove
from time import sleep
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Color, PatternFill, Font, Side, Border, Alignment
from constants import *
import calculate


class ExcelFile:

    def __init__(self):
        self.create_file(False)
        self.save()

    def create_file(self, new: bool):
        if new:
            if path.exists(path.abspath(RESULT_FILE_NAME)):
                remove(path.abspath(RESULT_FILE_NAME))
        elif path.exists(path.abspath(RESULT_FILE_NAME)):
            try:
                self.wb = load_workbook(path.abspath(RESULT_FILE_NAME))
                return
            except zipfile.BadZipFile:
                print('Файл поврежден, пересоздаю файл')

        self.wb = Workbook()
        for name in ('Архив', 'Отфильтрованные записи'):
            self.wb.create_sheet(name)

            for row in range(1, 8, 2):
                for column, (name_col, color) in enumerate(zip(tuple(DICT_OF_NAME_COLUMNS.get(row) + title for title in COLUMNS), COLORS), start=1):
                    cell = self.wb[name].cell(row, column, name_col)
                    cell.alignment = Alignment(wrapText=True, horizontal='center')
                    self.cell_decorate(cell, color, {'size': 10})

            for _ in range(4):
                self.wb[name].append(('',))

            self.wb[name].append(('Тираж', *(x for x in range(1, 13)), 'Сумма'))
        self.wb.remove(self.wb['Sheet'])

    def save(self):
        while True:
            try:
                self.wb.save(path.abspath(RESULT_FILE_NAME))
                # self.wb.close()
            except:
                sleep(1)
                continue
            else:
                return

    def write_data(self, data: tuple, table: str):
        """
        write to excel circulations
        :param table:
        :param data: number_of_circl + tuple a numbers
        :return:
        """
        if table == 'Архив' and len(data) == 2:
            data = data[0]
            if int(self.wb[table].cell(13, 1).value) < data[0][0]:
                last_code = int(self.wb[table].cell(13, 1).value)
                self.wb[table].insert_rows(13, data[0][0] - last_code)
                row, column = 13, 1
                data = data[:data[0][0] - last_code]
                for one_data in data:
                    for value in (one_data + (sum(one_data[1:]),)):
                        self.wb[table].cell(row, column, value)
                        column += 1
                    row += 1
                    column = 1
                self.read_old_static()
        else:
            for circulation in data:
                self.wb[table].append(circulation + (sum(circulation[1:]),))
        self.save()

    def write_statistics(self, data: tuple, table: str):
        calc = calculate.Calculator(data)
        for index, (first, second, third, fourth, color) in enumerate(zip(calc.first_dict_answer.values(),
                                                                          calc.second_dict_answer.values(),
                                                                          calc.third_dict_answer.values(),
                                                                          calc.fourth_dict_answer.values(),
                                                                          COLORS),
                                                                      start=1):

            for row, price in zip((2, 4, 6, 8), (first, second, third, fourth)):
                if not price:
                    continue
                self.cell_decorate(self.wb[table].cell(row, index, price), color)

        self.save()

    def cell_decorate(self, cell, color, Font_=FONT):
        cell.fill = PatternFill(patternType='solid', fgColor=Color(rgb=color))
        cell.font = Font(**Font_)
        cell.border = Border(**{key: Side(style=value) for key, value in BORDER.items()})

    def clear_filter(self):
        self.wb['Отфильтрованные записи'].delete_rows(13, 5000)
        self.save()

    def read_old_static(self):
        row = 13
        sums = []
        while self.wb['Архив'].cell(row, 14).value:
            sums.append(self.wb['Архив'].cell(row, 14).value)
            row += 1
        self.write_statistics(tuple(sums), 'Архив')

    def make_static(self, needed: int):
        row, column = 13, 1
        list_of_sum = []
        while needed and self.wb['Архив'].cell(row, column).value:
            sum_ = self.wb['Архив'].cell(row, column).value
            self.wb['Отфильтрованные записи'].cell(row, column, sum_)
            if column == 14:
                list_of_sum.append(sum_)
                column = 1
                row += 1
                needed -= 1
            else:
                column += 1
        self.write_statistics(tuple(list_of_sum), 'Отфильтрованные записи')

    def close(self):
        self.wb.close()


def main():
    pass


if __name__ == '__main__':
    main()


